import io
import openpyxl
from openpyxl.styles import Font


def users_to_excel(users) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    headers = ["tg_id", "full_name", "username", "is_blocked"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for u in users:
        ws.append([u.tg_id, u.full_name, u.username or "", "Ha" if u.is_blocked else "Yo'q"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def excel_to_users(data: bytes) -> list[dict]:
    wb = openpyxl.load_workbook(io.BytesIO(data))
    ws = wb.active
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        tg_id = row[0]
        full_name = row[1] if len(row) > 1 else None
        if not tg_id:
            continue
        try:
            tg_id = int(str(tg_id).strip())
        except (ValueError, TypeError):
            continue
        users.append({
            "tg_id": tg_id,
            "full_name": str(full_name).strip() if full_name else "",
            "username": None,
        })
    return users
